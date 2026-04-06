from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any, Mapping
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_TESTCASE_EXPORT_ROWS = 10000
TESTCASE_EXPORT_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_TESTCASE_DOCUMENT_EXPORT_ROWS = 10000
TESTCASE_DOCUMENT_EXPORT_MEDIA_TYPE = TESTCASE_EXPORT_MEDIA_TYPE

DEFAULT_TESTCASE_EXPORT_COLUMNS = [
    "case_id",
    "title",
    "module_name",
    "priority",
    "status",
    "description",
    "preconditions",
    "steps",
    "expected_results",
    "test_type",
    "design_technique",
    "test_data",
    "remarks",
    "quality_gate",
    "quality_score",
    "batch_id",
    "source_document_ids",
    "created_at",
    "updated_at",
]

TESTCASE_EXPORT_COLUMN_DEFS: dict[str, tuple[str, int]] = {
    "case_id": ("Case ID", 18),
    "title": ("标题", 28),
    "module_name": ("模块", 18),
    "priority": ("优先级", 10),
    "status": ("状态", 12),
    "description": ("描述", 28),
    "preconditions": ("前置条件", 20),
    "steps": ("步骤", 42),
    "expected_results": ("预期结果", 42),
    "test_type": ("测试类型", 14),
    "design_technique": ("设计技术", 16),
    "test_data": ("测试数据", 24),
    "remarks": ("备注", 18),
    "quality_gate": ("质量门禁", 12),
    "quality_score": ("质量分数", 12),
    "batch_id": ("批次 ID", 38),
    "source_document_ids": ("来源文档 IDs", 32),
    "created_at": ("创建时间", 22),
    "updated_at": ("更新时间", 22),
}


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_mapping(value: Any) -> dict[str, Any]:
    return dict(value) if isinstance(value, Mapping) else {}


def _coerce_list(value: Any) -> list[Any]:
    return list(value) if isinstance(value, list) else []


def _join_lines(value: Any) -> str:
    items: list[str] = []
    for item in _coerce_list(value):
        text = _coerce_text(item)
        if text:
            items.append(text)
    return "\n".join(items)


def _json_dump(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, indent=2, default=str)


def _content_json(case: Mapping[str, Any]) -> dict[str, Any]:
    return _coerce_mapping(case.get("content_json"))


def _content_meta(case: Mapping[str, Any]) -> dict[str, Any]:
    return _coerce_mapping(_content_json(case).get("meta"))


def _case_field(case: Mapping[str, Any], key: str) -> Any:
    if key in case and case.get(key) not in (None, ""):
        return case.get(key)
    return _content_json(case).get(key)


def _quality_review(case: Mapping[str, Any]) -> dict[str, Any]:
    return _coerce_mapping(_content_meta(case).get("quality_review"))


def _case_column_value(case: Mapping[str, Any], column: str) -> str:
    quality_review = _quality_review(case)
    mapping: dict[str, Any] = {
        "case_id": _coerce_text(_case_field(case, "case_id")),
        "title": _coerce_text(_case_field(case, "title")),
        "module_name": _coerce_text(_case_field(case, "module_name")),
        "priority": _coerce_text(_case_field(case, "priority")),
        "status": _coerce_text(_case_field(case, "status") or case.get("status")),
        "description": _coerce_text(_case_field(case, "description") or case.get("description")),
        "preconditions": _join_lines(_case_field(case, "preconditions")),
        "steps": _join_lines(_case_field(case, "steps")),
        "expected_results": _join_lines(_case_field(case, "expected_results")),
        "test_type": _coerce_text(_case_field(case, "test_type")),
        "design_technique": _coerce_text(_case_field(case, "design_technique")),
        "test_data": _json_dump(_case_field(case, "test_data")),
        "remarks": _coerce_text(_case_field(case, "remarks")),
        "quality_gate": _coerce_text(quality_review.get("quality_gate")),
        "quality_score": _coerce_text(quality_review.get("quality_score")),
        "batch_id": _coerce_text(case.get("batch_id")),
        "source_document_ids": _join_lines(
            case.get("source_document_ids") or _content_meta(case).get("source_document_ids")
        ),
        "created_at": _coerce_text(case.get("created_at")),
        "updated_at": _coerce_text(case.get("updated_at")),
    }
    return _coerce_text(mapping.get(column))


def resolve_testcase_export_columns(columns: list[str] | tuple[str, ...] | None) -> list[str]:
    if not columns:
        return list(DEFAULT_TESTCASE_EXPORT_COLUMNS)
    resolved: list[str] = []
    for item in columns:
        key = _coerce_text(item)
        if key and key in TESTCASE_EXPORT_COLUMN_DEFS and key not in resolved:
            resolved.append(key)
    return resolved or list(DEFAULT_TESTCASE_EXPORT_COLUMNS)


def build_case_export_content_disposition(filename: str) -> str:
    encoded = quote(filename)
    return f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded}"


def build_document_export_content_disposition(filename: str) -> str:
    encoded = quote(filename)
    return f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded}"


def _styled_row(
    sheet: Any,
    values: list[str | int],
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
) -> list[WriteOnlyCell]:
    row: list[WriteOnlyCell] = []
    for value in values:
        cell = WriteOnlyCell(sheet, value=value)
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if alignment is not None:
            cell.alignment = alignment
        row.append(cell)
    return row


def _case_export_filename(project_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    project_key = (project_id or "project")[:8]
    return f"testcase-cases-{project_key}-{timestamp}.xlsx"


def build_testcase_cases_workbook(
    *,
    project_id: str,
    cases: list[dict[str, Any]],
    filters: Mapping[str, Any],
    columns: list[str] | None = None,
) -> tuple[str, bytes]:
    resolved_columns = resolve_testcase_export_columns(columns)
    workbook = Workbook(write_only=True)
    cases_sheet = workbook.create_sheet("测试用例")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    widths = [TESTCASE_EXPORT_COLUMN_DEFS[column][1] for column in resolved_columns]
    for index, width in enumerate(widths, start=1):
        cases_sheet.column_dimensions[get_column_letter(index)].width = width
    cases_sheet.freeze_panes = "A2"
    cases_sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(resolved_columns))}{max(len(cases) + 1, 1)}"
    )

    headers = [TESTCASE_EXPORT_COLUMN_DEFS[column][0] for column in resolved_columns]
    cases_sheet.append(
        _styled_row(
            cases_sheet,
            headers,
            font=header_font,
            fill=header_fill,
            alignment=wrap_alignment,
        )
    )

    for case in cases:
        row = [_case_column_value(case, column) for column in resolved_columns]
        cases_sheet.append(_styled_row(cases_sheet, row, alignment=wrap_alignment))

    meta_sheet = workbook.create_sheet("导出信息")
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["字段", "值"],
            font=header_font,
            fill=header_fill,
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(_styled_row(meta_sheet, ["project_id", project_id], alignment=wrap_alignment))
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["导出时间", datetime.now().isoformat(timespec="seconds")],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(meta_sheet, ["导出总条数", str(len(cases))], alignment=wrap_alignment)
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["batch_id", _coerce_text(filters.get("batch_id"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["status", _coerce_text(filters.get("status"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["query", _coerce_text(filters.get("query"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.column_dimensions["A"].width = 18
    meta_sheet.column_dimensions["B"].width = 48

    output = BytesIO()
    workbook.save(output)
    return _case_export_filename(project_id), output.getvalue()


def _runtime_meta(document: Mapping[str, Any]) -> dict[str, Any]:
    provenance = _coerce_mapping(document.get("provenance"))
    return _coerce_mapping(provenance.get("runtime"))


def _asset_meta(document: Mapping[str, Any]) -> dict[str, Any]:
    provenance = _coerce_mapping(document.get("provenance"))
    return _coerce_mapping(provenance.get("asset"))


def _parsed_text_excerpt(document: Mapping[str, Any], *, max_chars: int = 8000) -> str:
    parsed_text = _coerce_text(document.get("parsed_text"))
    if not parsed_text:
        return ""
    return parsed_text[:max_chars]


def _document_export_filename(project_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    project_key = (project_id or "project")[:8]
    return f"testcase-documents-{project_key}-{timestamp}.xlsx"


def build_testcase_documents_workbook(
    *,
    project_id: str,
    documents: list[dict[str, Any]],
    filters: Mapping[str, Any],
) -> tuple[str, bytes]:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("PDF解析记录")
    headers = [
        "Document ID",
        "文件名",
        "content_type",
        "source_kind",
        "parse_status",
        "batch_id",
        "summary_for_model",
        "parsed_text_excerpt",
        "parsed_text_length",
        "structured_data_json",
        "thread_id",
        "run_id",
        "agent_key",
        "storage_path",
        "related_cases_count",
        "created_at",
    ]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        1: 38,
        2: 28,
        3: 22,
        4: 16,
        5: 14,
        6: 36,
        7: 42,
        8: 60,
        9: 16,
        10: 42,
        11: 28,
        12: 28,
        13: 18,
        14: 44,
        15: 16,
        16: 24,
    }
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(documents) + 1, 1)}"
    sheet.append(
        _styled_row(
            sheet,
            headers,
            font=header_font,
            fill=header_fill,
            alignment=wrap_alignment,
        )
    )

    for document in documents:
        runtime_meta = _runtime_meta(document)
        parsed_text = _coerce_text(document.get("parsed_text"))
        row = [
            _coerce_text(document.get("id")),
            _coerce_text(document.get("filename")),
            _coerce_text(document.get("content_type")),
            _coerce_text(document.get("source_kind")),
            _coerce_text(document.get("parse_status")),
            _coerce_text(document.get("batch_id")),
            _coerce_text(document.get("summary_for_model")),
            _parsed_text_excerpt(document),
            len(parsed_text),
            _json_dump(document.get("structured_data")),
            _coerce_text(runtime_meta.get("thread_id")),
            _coerce_text(runtime_meta.get("run_id")),
            _coerce_text(runtime_meta.get("agent_key")),
            _coerce_text(document.get("storage_path") or _asset_meta(document).get("storage_path")),
            _coerce_text(document.get("related_cases_count")),
            _coerce_text(document.get("created_at")),
        ]
        sheet.append(_styled_row(sheet, row, alignment=wrap_alignment))

    meta_sheet = workbook.create_sheet("导出信息")
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["字段", "值"],
            font=header_font,
            fill=header_fill,
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(_styled_row(meta_sheet, ["project_id", project_id], alignment=wrap_alignment))
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["导出时间", datetime.now().isoformat(timespec="seconds")],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(meta_sheet, ["导出总条数", str(len(documents))], alignment=wrap_alignment)
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["batch_id", _coerce_text(filters.get("batch_id"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["parse_status", _coerce_text(filters.get("parse_status"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["query", _coerce_text(filters.get("query"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.column_dimensions["A"].width = 18
    meta_sheet.column_dimensions["B"].width = 48

    output = BytesIO()
    workbook.save(output)
    return _document_export_filename(project_id), output.getvalue()
