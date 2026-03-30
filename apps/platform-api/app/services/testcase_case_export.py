from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any, Mapping
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_TESTCASE_EXPORT_ROWS = 2000
_EXPORT_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_mapping(value: Any) -> dict[str, Any]:
    return dict(value) if isinstance(value, Mapping) else {}


def _coerce_list(value: Any) -> list[Any]:
    return list(value) if isinstance(value, list) else []


def _join_lines(value: Any) -> str:
    items = []
    for item in _coerce_list(value):
        text = _coerce_text(item)
        if text:
            items.append(text)
    return "\n".join(items)


def _json_dump(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, indent=2, default=str)


def _content_json(case: Mapping[str, Any]) -> dict[str, Any]:
    return _coerce_mapping(case.get("content_json"))


def _content_meta(case: Mapping[str, Any]) -> dict[str, Any]:
    return _coerce_mapping(_content_json(case).get("meta"))


def _field(case: Mapping[str, Any], key: str) -> Any:
    if key in case and case.get(key) not in (None, ""):
        return case.get(key)
    return _content_json(case).get(key)


def _export_filename(project_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    project_key = (project_id or "project")[:8]
    return f"testcase-cases-{project_key}-{timestamp}.xlsx"


def build_content_disposition(filename: str) -> str:
    encoded = quote(filename)
    return f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded}"


def build_testcase_cases_workbook(
    *,
    project_id: str,
    cases: list[dict[str, Any]],
    filters: Mapping[str, Any],
) -> tuple[str, bytes]:
    workbook = Workbook()
    cases_sheet = workbook.active
    cases_sheet.title = "测试用例"

    headers = [
        "Case ID",
        "标题",
        "模块",
        "优先级",
        "状态",
        "描述",
        "前置条件",
        "步骤",
        "预期结果",
        "测试类型",
        "设计技术",
        "测试数据",
        "备注",
        "质量门禁",
        "质量分数",
        "批次 ID",
        "来源文档 IDs",
        "创建时间",
        "更新时间",
    ]
    cases_sheet.append(headers)

    for case in cases:
        meta = _content_meta(case)
        quality_review = _coerce_mapping(meta.get("quality_review"))
        row = [
            _coerce_text(_field(case, "case_id")),
            _coerce_text(_field(case, "title")),
            _coerce_text(_field(case, "module_name")),
            _coerce_text(_field(case, "priority")),
            _coerce_text(_field(case, "status") or case.get("status")),
            _coerce_text(_field(case, "description") or case.get("description")),
            _join_lines(_field(case, "preconditions")),
            _join_lines(_field(case, "steps")),
            _join_lines(_field(case, "expected_results")),
            _coerce_text(_field(case, "test_type")),
            _coerce_text(_field(case, "design_technique")),
            _json_dump(_field(case, "test_data")),
            _coerce_text(_field(case, "remarks")),
            _coerce_text(quality_review.get("quality_gate")),
            _coerce_text(quality_review.get("quality_score")),
            _coerce_text(case.get("batch_id")),
            _join_lines(case.get("source_document_ids") or meta.get("source_document_ids")),
            _coerce_text(case.get("created_at")),
            _coerce_text(case.get("updated_at")),
        ]
        cases_sheet.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    for cell in cases_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment

    for row in cases_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    widths = {
        1: 18,
        2: 28,
        3: 18,
        4: 10,
        5: 12,
        6: 28,
        7: 20,
        8: 42,
        9: 42,
        10: 14,
        11: 16,
        12: 24,
        13: 18,
        14: 12,
        15: 12,
        16: 38,
        17: 32,
        18: 22,
        19: 22,
    }
    for index, width in widths.items():
        cases_sheet.column_dimensions[get_column_letter(index)].width = width
    cases_sheet.freeze_panes = "A2"
    cases_sheet.auto_filter.ref = cases_sheet.dimensions

    meta_sheet = workbook.create_sheet("导出信息")
    meta_sheet.append(["字段", "值"])
    meta_sheet.append(["project_id", project_id])
    meta_sheet.append(["导出时间", datetime.now().isoformat(timespec="seconds")])
    meta_sheet.append(["导出总条数", len(cases)])
    meta_sheet.append(["batch_id", _coerce_text(filters.get("batch_id"))])
    meta_sheet.append(["status", _coerce_text(filters.get("status"))])
    meta_sheet.append(["query", _coerce_text(filters.get("query"))])
    for cell in meta_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    meta_sheet.column_dimensions["A"].width = 18
    meta_sheet.column_dimensions["B"].width = 48
    for row in meta_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    output = BytesIO()
    workbook.save(output)
    return _export_filename(project_id), output.getvalue()


__all__ = [
    "MAX_TESTCASE_EXPORT_ROWS",
    "_EXPORT_MEDIA_TYPE",
    "build_content_disposition",
    "build_testcase_cases_workbook",
]
